"""
SheetCopilot v2: Enhanced multi-stage reasoning for real-world spreadsheets
Specifically designed for SpreadsheetBench's two key characteristics:
1. Complex Instructions from Real World - Natural language understanding
2. Spreadsheet in Diverse Formats - Non-standard tables, multiple tables, rich formats

Architecture: Observing → Understanding → Planning → Implementing → Validating → Executing
"""
# 修改日志记录脚本，只保留最关键对优化模型有用的信息，比如生成的脚本与测试记录，保持记录和代码的简洁
import os
import re
import json
import logging
import argparse
from tqdm import tqdm
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple

from llm_api import get_llm_response
from code_exec import get_exec_client, extract_code, exec_code
from excel_calculator import calculate_formulas
from excel_recalc import recalc_workbook


def setup_logger(log_dir: str, model_name: str, dataset_name: str = None) -> logging.Logger:
    """Setup comprehensive logging system"""
    os.makedirs(log_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    # Format: timestamp_dataset_model.log
    dataset_part = f"_{dataset_name}" if dataset_name else ""
    log_file = f"{log_dir}/{timestamp}{dataset_part}_{model_name}.log"
    
    logger = logging.getLogger('SheetCopilot_v2')
    logger.setLevel(logging.DEBUG)
    
    fh = logging.FileHandler(log_file, encoding='utf-8')
    fh.setLevel(logging.DEBUG)
    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    
    formatter = logging.Formatter(
        '[%(asctime)s] [%(levelname)s] [%(funcName)s]\n%(message)s\n',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    fh.setFormatter(formatter)
    ch.setFormatter(logging.Formatter('[%(levelname)s] %(message)s'))
    
    logger.addHandler(fh)
    logger.addHandler(ch)
    return logger


class SheetCopilotV2:
    """
    Enhanced SheetCopilot with improved understanding of:
    - Non-standard table layouts (data not starting at A1)
    - Multiple tables in single/multiple sheets
    - Complex natural language instructions from real users
    - Rich formatting and non-textual elements
    """
    
    def __init__(self, opt, logger: logging.Logger):
        self.opt = opt
        self.logger = logger
        self.client = get_exec_client(opt.code_exec_url, opt.conv_id)
        self.stage_history = []
        self.max_revisions = getattr(opt, 'max_revisions', 3)
        self.enable_timing = getattr(opt, 'enable_timing', True)  # 控制是否启用计时
        self.debug = getattr(opt, 'debug', False)  # 控制是否输出详细调试信息
        self.stage_timings = {}  # 存储每个阶段的运行时间
        
    def log_stage(self, stage: str, content: str, stage_time: float = None):
        """Track stage execution (minimal logging)"""
        stage_record = {
            'stage': stage,
            'content': content[:200],  # Truncate for storage
            'timestamp': datetime.now().isoformat()
        }
        if self.enable_timing and stage_time is not None:
            stage_record['duration'] = stage_time
        self.stage_history.append(stage_record)
    
    def stage_1_deep_observation(self, file_path: str, instruction: str, 
                                 answer_position: str, instruction_type: str) -> Dict[str, Any]:
        """
        Stage 1: Deep Observation - Understanding NON-STANDARD spreadsheet structures
        
        Key focus areas for diverse formats:
        1. Identify actual data region (not assuming A1 start)
        2. Detect multiple tables in single sheet
        3. Find headers and their positions
        4. Detect merged cells, empty regions, formatting
        5. Understand multi-sheet relationships
        
        NOTE: This stage uses pre-defined observation code instead of LLM generation
        to ensure stability and avoid common errors (e.g., incorrect imports).
        """
        import time
        stage_start = time.time() if self.enable_timing else None
        
        if self.debug:
            print(f"\n[DEBUG] 🚀 Starting Stage 1: Deep Observation")
        
        # Pre-defined observation code - no LLM generation needed
        # Use triple quotes without f-string to avoid escaping issues
        observation_code = """import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries
import re

# Phase 1: Global Structure Analysis
wb = openpyxl.load_workbook('""" + file_path + """')

print("📊 WORKBOOK STRUCTURE:")
print(f"All sheets: {wb.sheetnames}")
print(f"Active sheet: {wb.active.title}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\\n--- Sheet: {sheet_name} ---")
    print(f"Dimensions: {ws.max_row} rows × {ws.max_column} cols")
    
    # Find actual data boundaries
    min_row, max_row = None, None
    min_col, max_col = None, None
    for row in range(1, ws.max_row + 1):
        if any(ws.cell(row, col).value is not None for col in range(1, ws.max_column + 1)):
            if min_row is None:
                min_row = row
            max_row = row
    for col in range(1, ws.max_column + 1):
        if any(ws.cell(row, col).value is not None for row in range(1, ws.max_row + 1)):
            if min_col is None:
                min_col = col
            max_col = col
    
    if min_row and min_col:
        print(f"Actual data region: Row {min_row}-{max_row}, Col {min_col}-{max_col}")
        print(f"Column letters: {get_column_letter(min_col)}-{get_column_letter(max_col)}")

# Phase 2: Target Position Analysis
target_str = \"""" + answer_position + """\"
sheet_match = re.match(r"'([^']+)'!(.+)", target_str)
if sheet_match:
    target_sheet = sheet_match.group(1)
    target_range = sheet_match.group(2)
    print(f"\\n🎯 TARGET: Sheet '{target_sheet}', Range '{target_range}'")
    ws = wb[target_sheet]
else:
    target_range = target_str
    ws = wb.active
    print(f"\\n🎯 TARGET: Active sheet, Range '{target_range}'")

print(f"\\n📍 TARGET CELL ANALYSIS:")
try:
    min_col, min_row, max_col, max_row = range_boundaries(target_range)
    print(f"Target range: {target_range}, min_row={min_row}, max_row={max_row}, min_col={min_col}, max_col={max_col}")
    total_rows = max_row - min_row + 1
    # For large ranges (>20 rows), show sample only
    if total_rows > 20:
        print(f"Large range detected ({total_rows} rows). Showing first 10 and last 5 rows as sample:")
        sample_rows = list(range(min_row, min(min_row + 10, max_row + 1))) + list(range(max(max_row - 4, min_row + 10), max_row + 1))
    else:
        sample_rows = range(min_row, max_row + 1)
    
    for row in sample_rows:
        values = []
        coords = []
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            values.append(cell.value)
            coords.append(cell.coordinate)
        print(f"Row {row}: {coords} = {values}")
    
    if total_rows > 20:
        print(f"... ({total_rows - 15} middle rows omitted) ...")
except Exception as e:
    print(f"⚠️ Could not analyze target range in detail: {str(e)}")
    print(f"Attempting to access as single cell or use fallback...")
    try:
        cell = ws[target_range]
        print(f"Single cell {cell.coordinate} = {cell.value}")
    except:
        print(f"Target range is complex. Will handle dynamically in code.")

# Phase 3: Context & Merged Cells
print(f"\\n🔗 MERGED CELLS:")
merged_ranges = ws.merged_cells.ranges
for merged in merged_ranges:
    print(f"Merged: {str(merged)}")

# Phase 4: Answer Position Format Analysis (as reference example)
print(f"\\n📝 ANSWER POSITION CURRENT CONTENT (Format Reference):")
try:
    # Read current content in answer position to understand expected format
    answer_content = []
    answer_has_data = False
    min_col, min_row, max_col, max_row = range_boundaries(target_range)
    
    # Sample up to 10 cells to understand format
    sample_limit = min(10, (max_row - min_row + 1) * (max_col - min_col + 1))
    cell_count = 0
    
    for row in range(min_row, max_row + 1):
        if cell_count >= sample_limit:
            break
        for col in range(min_col, max_col + 1):
            if cell_count >= sample_limit:
                break
            cell = ws.cell(row=row, column=col)
            cell_value = cell.value
            cell_type = type(cell_value).__name__
            cell_format = cell.number_format if hasattr(cell, 'number_format') else 'General'
            
            # Check if cell has formula (correct detection)
            if hasattr(cell, 'data_type') and cell.data_type == 'f':
                formula = cell.value  # This IS the formula string
                print(f"  {cell.coordinate}: FORMULA = {formula}")
            else:
                print(f"  {cell.coordinate}: VALUE = {cell_value} (type={cell_type}, format={cell_format})")
            
            if cell_value is not None and cell_value != '':
                answer_has_data = True
                answer_content.append({
                    'coord': cell.coordinate,
                    'value': str(cell_value)[:50],  # Truncate long values
                    'type': cell_type,
                    'format': cell_format
                })
            cell_count += 1
    
    if answer_has_data:
        print(f"\\n✅ Answer position contains data - USE AS FORMAT REFERENCE")
        print(f"   Total non-empty cells sampled: {len(answer_content)}")
        
        # Check if answer cells contain formulas
        formula_cells = []
        for row in range(min_row, min(min_row + 5, max_row + 1)):  # Check first 5 rows
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                if hasattr(cell, 'data_type') and cell.data_type == 'f':
                    formula_cells.append((cell.coordinate, cell.value))
        
        if formula_cells:
            print(f"\\n   🎯 FORMULA PATTERN DETECTED:")
            print(f"   - {len(formula_cells)} cells contain formulas")
            print(f"   - First formula example: {formula_cells[0][1]}")
            if len(formula_cells) > 1:
                print(f"   - Second formula example: {formula_cells[1][1]}")
            print(f"   ⚠️ CRITICAL: Must generate FORMULAS, not static values!")
            print(f"   💡 Analyze the formula pattern and apply it with correct row references")
            
            # NEW: Extract referenced cells/ranges from formulas to understand data dependencies
            print(f"\\n   📊 DATA TYPE & DEPENDENCY ANALYSIS:")
            referenced_cells = set()
            for coord, formula in formula_cells[:3]:  # Analyze first 3 formulas
                # Extract cell references (A1, $B$2, etc.)
                import re
                refs = re.findall(r'\\$?[A-Z]+\\$?[0-9]+', formula)
                referenced_cells.update(refs)
            
            if referenced_cells:
                print(f"   - Referenced cells in formulas: {', '.join(sorted(referenced_cells)[:10])}")
                print(f"\\n   📋 DATA TYPE OF REFERENCED CELLS:")
                for ref in sorted(referenced_cells)[:5]:  # Show first 5
                    try:
                        ref_clean = ref.replace('$', '')
                        ref_cell = ws[ref_clean]
                        ref_value = ref_cell.value
                        ref_type = type(ref_value).__name__
                        ref_format = ref_cell.number_format if hasattr(ref_cell, 'number_format') else 'General'
                        print(f"   - {ref_clean}: {ref_value} (type={ref_type}, format={ref_format})")
                    except:
                        pass
        else:
            print(f"   Format patterns detected:")
            # Summarize types
            types_found = set(item['type'] for item in answer_content)
            formats_found = set(item['format'] for item in answer_content if item['format'] != 'General')
            print(f"   - Data types: {', '.join(types_found)}")
            if formats_found:
                print(f"   - Number formats: {', '.join(formats_found)}")
            
            # NEW: Answer Source Analysis - discover where answer values come from
            print(f"\\n   🔍 ANSWER SOURCE ANALYSIS (comparing with nearby columns):")
            # Get answer column letter
            answer_col_idx = min_col
            from openpyxl.utils import get_column_letter
            answer_col_letter = get_column_letter(answer_col_idx)
            
            # Check 3 rows before answer column (likely input columns)
            nearby_cols = range(max(1, answer_col_idx - 3), answer_col_idx)
            sample_rows_for_analysis = list(range(min_row, min(min_row + 5, max_row + 1)))
            
            print(f"   Comparing answer column {answer_col_letter} with nearby columns:")
            for row_idx in sample_rows_for_analysis:
                answer_val = ws.cell(row=row_idx, column=answer_col_idx).value
                if answer_val in (None, ""):
                    continue
                    
                row_info = f"   Row {row_idx}: Answer={answer_val}"
                
                # Check each nearby column to find potential source
                for col_idx in nearby_cols:
                    col_letter = get_column_letter(col_idx)
                    col_val = ws.cell(row=row_idx, column=col_idx).value
                    
                    # Check if answer is substring or exact match
                    if col_val and str(answer_val) in str(col_val):
                        row_info += f" | {col_letter}={col_val}"
                        if str(col_val) == str(answer_val):
                            row_info += " ✓EXACT"
                        else:
                            row_info += " ✓CONTAINS"
                
                print(row_info)
            
            print(f"\\n   💡 INSIGHT: Analyze which column contains the answer values!")
            print(f"   - If answer is EXACT match of a column → Simple copy")
            print(f"   - If answer is SUBSTRING of a column → Extract/parse logic needed")
            print(f"   - Check if other columns act as INDEX/CONDITION for extraction")
            
            # NEW: For non-formula cells, analyze actual data type patterns
            print(f"\\n   📊 REFERENCE DATA TYPE ANALYSIS:")
            for item in answer_content[:5]:  # Show first 5 cells
                print(f"   - {item['coord']}: type={item['type']}, value_sample={item['value'][:30]}")
    else:
        print(f"⚠️ Answer position is empty - no format reference available")
        
except Exception as e:
    print(f"⚠️ Could not analyze answer position format: {str(e)}")

# Phase 5: Pattern Recognition
print("\\n🎯 TASK PATTERN RECOGNITION:")
# Instruction and type are shown in logs, no need to print in code

wb.close()
"""
        
        # Execute the observation code (no logging to reduce verbosity)
        try:
            result = exec_code(self.client, observation_code)
            # Check for fatal errors only (warnings ⚠️ are OK)
            has_fatal_error = 'Traceback' in result or 'JSON_DECODE_ERROR' in result or 'EXECUTION REQUEST ERROR' in result
            
            # Check if result is actually SOURCE CODE (not executed)
            is_source_code = ('import openpyxl' in result and 'wb = openpyxl.load_workbook' in result and result.count('\n') < 5)
            
            # Check if we got minimal required info from execution
            has_basic_info = ('Target range:' in result or 'WORKBOOK STRUCTURE:' in result or 'All sheets:' in result)
            
            # Success: no fatal error, not source code, has basic info
            success = (not has_fatal_error) and (not is_source_code) and has_basic_info
            
            # Only log critical errors
        except Exception as e:
            result = f"Observation error: {str(e)}"
            success = False
        
        # Create a summary prompt for context (used in later stages)
        observation_summary = f"""📊 SPREADSHEET OBSERVATION COMPLETED

🎯 Task: {instruction}
📋 Type: {instruction_type}
🎯 Target: {answer_position}
📂 File: {file_path}

Observation Results:
{result}
"""

        if self.enable_timing:
            stage_time = time.time() - stage_start
            self.stage_timings['stage_1_observation'] = stage_time
            if self.debug:
                print(f"[DEBUG] ✅ Stage 1 completed in {stage_time:.2f}s")
            self.log_stage("1️⃣ DEEP OBSERVATION", 
                          f"Analyzing non-standard spreadsheet structure\nTask: {instruction}", 
                          stage_time)

        return {
            'prompt': observation_summary,
            'response': result,
            'code': observation_code,
            'result': result,
            'success': success
        }
    
    def stage_2_instruction_understanding(self, observation: Dict[str, Any], 
                                         instruction: str, instruction_type: str) -> Dict[str, Any]:
        """
        Stage 2: Instruction Understanding - Parse COMPLEX REAL-WORLD natural language
        
        Real-world instructions characteristics:
        - Long, descriptive, informal language
        - May contain background context and explanations
        - Multiple requirements in one sentence
        - Implicit assumptions and domain knowledge
        - References to attached files and examples
        """
        import time
        stage_start = time.time() if self.enable_timing else None
        
        if self.debug:
            print(f"\n[DEBUG] 🚀 Starting Stage 2: Instruction Understanding")
        
        understanding_prompt = f"""You are SheetCopilot v2 in INSTRUCTION UNDERSTANDING stage.


This is a REAL-WORLD user question from Excel forums. Your task is to extract the CORE requirements.

📝 **ORIGINAL INSTRUCTION** (may be long and informal):
{instruction}

📊 **SPREADSHEET STRUCTURE** (from observation):
{observation['result'][:1200]}  # Truncate for LLM prompt (increased to include format reference)

🎯 **TASK TYPE**: {instruction_type}

💡 **IMPORTANT**: Check if "ANSWER POSITION CURRENT CONTENT" section shows existing data - if yes, this is a FORMAT REFERENCE showing the expected output format (data type, number format, formula style, etc.). Your solution MUST preserve this format.

**YOUR ANALYSIS TASK**:
Break down this real-world instruction into structured requirements:

## 1. Core Objective
What is the PRIMARY goal? (in one clear sentence)

## 2. Input Data Location
- Which cells/ranges contain the INPUT data?
- Are there multiple source locations?
- What format is the input data? (numbers, text, formulas, etc.)

## 3. Output Requirements
- Where should results be written? (target cells)
- What format should output be? (formula, value, formatting, etc.)
- Any specific output constraints?

## 4. Business Logic
- What calculation/operation is needed?
- Any conditions or criteria to apply?
- Special cases or edge cases mentioned?

Provide your structured analysis:
"""
        
        # Generate understanding (logging disabled for brevity)
        response = get_llm_response(
            [observation['prompt'], observation['response'], understanding_prompt], 
            self.opt
        )
        
        if self.enable_timing:
            stage_time = time.time() - stage_start
            self.stage_timings['stage_2_understanding'] = stage_time
            if self.debug:
                print(f"[DEBUG] ✅ Stage 2 completed in {stage_time:.2f}s")
            self.log_stage("2️⃣ INSTRUCTION UNDERSTANDING", 
                          "Parsing complex natural language from real user", 
                          stage_time)
        
        return {
            'prompt': understanding_prompt,
            'response': response,
            'structured_requirements': self._parse_requirements(response)
        }
    
    def stage_3_solution_planning(self, observation: Dict[str, Any], 
                                 understanding: Dict[str, Any],
                                 file_path: str, output_path: str,
                                 answer_position: str) -> Dict[str, Any]:
        """
        Stage 3: Solution Planning - Design robust solution for non-standard formats
        
        Planning must account for:
        - Dynamic cell references (not hardcoded A1)
        - Multiple table navigation
        - Empty cell handling
        - Format preservation
        - Edge cases from real-world data
        """
        import time
        stage_start = time.time() if self.enable_timing else None
        
        if self.debug:
            print(f"\n[DEBUG] 🚀 Starting Stage 3: Solution Planning")
        
        planning_prompt = f"""You are SheetCopilot v2 in SOLUTION PLANNING stage.


📊 **SPREADSHEET FACTS** (non-standard structure):
{observation['result'][:1000]}  # Truncated (increased to include format reference)

🎯 **UNDERSTOOD REQUIREMENTS**:
{understanding['response'][:800]}  # Truncated

📂 **FILE PATHS**:
- Input: {file_path}
- Output: {output_path}
- Target cells: {answer_position}

💡 **FORMAT REFERENCE**: If observation shows existing data in answer position, PRESERVE that format (data type, number format, formula vs value). This is critical for correctness!

**YOUR PLANNING TASK**:
Design a step-by-step implementation plan that handles NON-STANDARD spreadsheet formats.

## Implementation Plan Template:

### Step 1: Load and Validate
```
- Load workbook from {file_path}
- Identify target sheet (handle multi-sheet case)
- Validate target range {answer_position} exists
- Check for merged cells or formatting in target area
```

### Step 2: Locate Input Data (DYNAMIC, not hardcoded!)
```
- Based on observation, input data is at: [SPECIFY ACTUAL LOCATION]
- NOT assuming A1 start!
- Handle empty cells: [STRATEGY]
- Account for non-standard table boundaries
```

### Step 3: Extract and Process
```
- Read input data using dynamic references
- Data type conversions needed: [SPECIFY]
- Handle edge cases: empty cells, merged cells, formulas vs values
- Validation checks before processing
```

### Step 4: Apply Business Logic
```
- Core operation: [DESCRIBE CLEARLY]
- Formula structure (if applicable): [FORMULA]
- Calculation steps: [ENUMERATE]
- Condition handling: [IF ANY]
```

### Step 5: Write Results
```
- Target cells: {answer_position}
- Write as: [FORMULA or VALUE or FORMATTED_VALUE]
- Preserve existing formatting: [YES/NO]
- Handle multiple target cells: [STRATEGY]
```

### Step 6: Save and Verify
```
- Save to {output_path}
- Verify write succeeded
- Close workbook properly
```

## Risk Mitigation:
- ❌ AVOID: Hardcoding cell references like A1, B2
- ✅ USE: Dynamic references based on observation results
- ❌ AVOID: Assuming headers in row 1
- ✅ USE: Actual header locations from analysis
- ❌ AVOID: Ignoring empty cells
- ✅ USE: Explicit null/empty checks

Provide your COMPLETE plan with SPECIFIC cell references based on the observation:
"""
        
        # Generate plan (logging disabled for brevity)
        messages = [
            observation['prompt'], observation['response'],
            understanding['prompt'], understanding['response'],
            planning_prompt
        ]
        response = get_llm_response(messages, self.opt)
        
        if self.enable_timing:
            stage_time = time.time() - stage_start
            self.stage_timings['stage_3_planning'] = stage_time
            if self.debug:
                print(f"[DEBUG] ✅ Stage 3 completed in {stage_time:.2f}s")
            self.log_stage("3️⃣ SOLUTION PLANNING", 
                          "Designing robust implementation for non-standard structure", 
                          stage_time)
        
        return {
            'prompt': planning_prompt,
            'response': response,
            'plan_steps': self._extract_steps(response)
        }
    
    def stage_4_code_implementation(self, observation: Dict[str, Any],
                                   understanding: Dict[str, Any],
                                   plan: Dict[str, Any],
                                   file_path: str, output_path: str,
                                   answer_position: str) -> Dict[str, Any]:
        """
        Stage 4: Code Implementation - Generate robust Python code
        
        Implementation principles:
        - Use dynamic cell references from observation
        - Handle all edge cases identified in planning
        - Include comprehensive error handling
        - Support diverse table formats
        """
        import time
        stage_start = time.time() if self.enable_timing else None
        
        if self.debug:
            print(f"\n[DEBUG] 🚀 Starting Stage 4: Code Implementation")
        
        implementation_prompt = f"""You are SheetCopilot v2 in CODE IMPLEMENTATION stage.


📊 **OBSERVED STRUCTURE**:
{observation['result'][:1000]}  # Truncated (increased to include format reference)

🎯 **REQUIREMENTS SUMMARY**:
{understanding['response'][:800]}

📋 **IMPLEMENTATION PLAN**:
{plan['response']}

**YOUR CODING TASK**:
Write COMPLETE, PRODUCTION-READY Python code following the plan above.

**🎯 FORMAT & DATA TYPE PRESERVATION (CRITICAL)**:
✅ Check observation's "ANSWER POSITION CURRENT CONTENT" section
✅ If it shows existing data, this is your FORMAT REFERENCE - match it exactly:
   
   **FORMULAS vs VALUES (MOST CRITICAL)**:
   - If shows "FORMULA PATTERN DETECTED" → MUST write FORMULA STRINGS (not computed values)
   - If shows "FORMULA = ..." → This is a STRING starting with "=", NOT a computed value
   - Writing formulas: `cell.value = '=IF(A1>0, "Yes", "No")'`  ← String literal
   - Writing values: `cell.value = "Yes"` or `cell.value = 123`  ← Plain data
   
   **DATA TYPE ANALYSIS**:
   - Observation shows "DATA TYPE OF REFERENCED CELLS" → use this to understand input types
   - If referenced cells are strings → formula should handle strings
   - If referenced cells are numbers → formula should handle numeric operations
   - Match the data type pattern observed in input answer column
   
   **FORMULA PATTERN REPLICATION**:
   - Example: If reference shows "FORMULA = =IF(E2="","", TRIM(LEFT(E2,...)))", your code should:
     1. Analyze: Formula references E2 (same row), checks if empty, then extracts substring
     2. Generate similar formula for each target row with RELATIVE references:
        * Row 2: `=IF(E2="","", TRIM(LEFT(E2,...)))`
        * Row 3: `=IF(E3="","", TRIM(LEFT(E3,...)))`
     3. Write formula STRING to cell.value: `ws.cell(row=2, column=7).value = '=IF(E2=...)'`
   
   **PLAIN VALUE PATTERN**:
   - Example: If reference shows "VALUE = 123 (type=int)" → write computed integers
   - Example: If reference shows "VALUE = 'text' (type=str)" → write computed strings

**CRITICAL REQUIREMENTS**:
✅ Use openpyxl library (already installed in Docker environment)
✅ NO hardcoded cell references - use DYNAMIC references from observation
✅ Handle empty cells explicitly (check `if cell.value is not None`)
✅ Include try-except for robust error handling
✅ Use actual sheet names and cell ranges from observation
✅ Support non-standard table positions
✅ Load from: {file_path}
✅ Save to: {output_path}
✅ Target cells: {answer_position}

🚫 **PROHIBITIONS (STRUCTURAL SAFETY)**:
❌ Do NOT create or copy to a temporary "helper" column and then delete it (e.g. copying G to H and calling `ws.delete_cols`).
❌ Do NOT call `ws.delete_cols()` unless the task explicitly requires column removal from the dataset semantics.
❌ Do NOT insert phantom columns just to preserve original values for a later formula.
✅ If you need the original cell value before overwriting, read it into a Python variable (e.g. `orig_val = ws.cell(row, col).value`) and then assign the final computed value back.
✅ Prefer computing final results in Python and writing plain values (avoid volatile formulas unless the instruction explicitly demands formulas as output).
✅ Avoid patterns that reference a column that will be deleted in the same script.

⚠️ **CRITICAL: AVOID CIRCULAR REFERENCES!**
❌ DO NOT write formulas that reference the target cell itself
❌ DO NOT create circular dependencies between target cells
❌ Example: If target is H3, do NOT use H3 in the formula
✅ Only reference INPUT data cells, never OUTPUT target cells

⚠️ **EXCEL FORMULA SYNTAX RULES** (when writing formulas to cells):
❌ WRONG: =@XLOOKUP(...) or @Sheet1!A1    → NO @ prefix before function names or sheet names
✅ CORRECT: =XLOOKUP(...) or Sheet1!A1

❌ WRONG: ="*&A1&*"                        → String literal cannot contain & without quotes
✅ CORRECT: ="*"&A1&"*"                    → Concatenate with & outside quotes

❌ WRONG: =IF(@B:B="value",A:A,"")         → NO @ prefix in array formulas
✅ CORRECT: =IF(B:B="value",A:A,"")        → Clean array formula syntax

When writing Excel formulas in Python code:
```python
# Correct string concatenation in formulas
cell.value = '="*"&A1&"*"'              # NOT '="*&A1&*"'
cell.value = '=XLOOKUP("*"&A1&"*",...)'  # NOT '=@XLOOKUP("*&A1&*",...)'
```

**CODE TEMPLATE** (adapt to your specific task):
```python
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import re

try:
    # 1. Load workbook
    print("Loading workbook...")
    wb = openpyxl.load_workbook('{file_path}')
    
    # 2. Get target sheet (handle sheet name in answer_position)
    target_str = "{answer_position}"
    sheet_match = re.match(r"'([^']+)'!(.+)", target_str)
    if sheet_match:
        sheet_name = sheet_match.group(1)
        target_range = sheet_match.group(2)
        ws = wb[sheet_name]
        print(f"Working on sheet: {{sheet_name}}")
    else:
        ws = wb.active
        target_range = target_str
    
    # 3. Parse target range (e.g., "A1:B10" or "C5")
    # Implement based on your plan
    
    # 4. Locate input data (DYNAMIC - from observation!)
    # Based on observation results, input data is at: [FILL FROM OBSERVATION]
    
    # 5. Read input data with null checks
    # for cell in ws[...]:
    #     if cell.value is not None:
    #         ...
    
    # 6. Process data (implement business logic)
    # [YOUR CORE LOGIC HERE]
    
    # 7. Write results to target cells
    # Handle both single cell and range cases
    
    # 8. Save output
    wb.save('{output_path}')
    wb.close()
    print(f"✅ Successfully saved to {output_path}")
    
except Exception as e:
    print(f"❌ Error: {{str(e)}}")
    import traceback
    traceback.print_exc()
```

**Generate COMPLETE implementation code now:**
"""
        
        # Generate implementation (prompt logging disabled)
        messages = [
            observation['prompt'], observation['response'],
            understanding['prompt'], understanding['response'],
            plan['prompt'], plan['response'],
            implementation_prompt
        ]
        response = get_llm_response(messages, self.opt)
        
        code = extract_code(response)
        # Replace placeholders with actual paths for current test case
        code = code.replace('{file_path}', file_path)
        code = code.replace('{output_path}', output_path)
        # Log generated code (KEY INFO for debugging)
        self.logger.info(f"[GENERATED CODE]\n{code}")
        
        if self.enable_timing:
            stage_time = time.time() - stage_start
            self.stage_timings['stage_4_implementation'] = stage_time
            if self.debug:
                print(f"[DEBUG] ✅ Stage 4 completed in {stage_time:.2f}s")
            self.log_stage("4️⃣ CODE IMPLEMENTATION", 
                          "Generating production-ready Python code", 
                          stage_time)
        
        return {
            'prompt': implementation_prompt,
            'response': response,
            'code': code
        }
    
    def stage_5_code_validation(self, implementation: Dict[str, Any],
                               observation: Dict[str, Any],
                               plan: Dict[str, Any],
                               file_path: str, output_path: str,
                               instruction: str, answer_position: str) -> Dict[str, Any]:
        """
        Stage 5: Code Validation - Execute and verify results
        
        NEW APPROACH:
        1. Execute the generated code first
        2. If successful, read the answer_position content from output
        3. Let AI judge if results are reasonable given the instruction
        4. If unreasonable or errors, provide corrected code
        """
        import time
        import openpyxl
        import re
        stage_start = time.time() if self.enable_timing else None
        
        if self.debug:
            print(f"\n[DEBUG] 🚀 Starting Stage 5: Code Validation (Execute & Verify)")
        
        # Replace paths before execution
        code_to_execute = self._replace_hardcoded_paths(implementation['code'], file_path, output_path)
        
        # Execute code for validation
        try:
            exec_result = exec_code(self.client, code_to_execute)
            has_error = 'Error' in exec_result or 'Traceback' in exec_result or '❌' in exec_result
            # Only log if there's an error
            if has_error:
                self.logger.info(f"[VALIDATION ERROR]\n{exec_result}")
            
            if has_error:
                # Execution failed - do static validation only
                validation_prompt = f"""You are SheetCopilot v2 in CODE VALIDATION stage.

The code execution FAILED. Please identify and fix the errors.

📋 **TASK**: {instruction}

📊 **OBSERVED DATA**: 
{observation['response'][:500]}

📋 **IMPLEMENTATION PLAN**:
{plan['response'][:600]}

💻 **GENERATED CODE** (has errors):
```python
{implementation['code']}
```

❌ **EXECUTION ERROR**:
```
{exec_result}
```

**YOUR TASK**:
1. Analyze the error message and traceback
2. Identify the root cause (e.g., hardcoded references, None values, index errors)
3. Provide CORRECTED code that fixes all issues

Provide the corrected code:
"""
                
            else:
                # Execution succeeded - verify result content and build rich feedback
                output_local = output_path.replace('/mnt/data/', '../data/')
                input_local = file_path.replace('/mnt/data/', '../data/')
                answer_content = "Could not read output file"
                input_answer_content = "Could not read input file"
                answer_values_raw = []
                input_answer_values_raw = []
                summary_json = {}
                input_summary_json = {}
                neighbor_alert = None
                
                # FIRST: Read input file's answer column to detect pattern
                try:
                    if os.path.exists(input_local):
                        wb_input = openpyxl.load_workbook(input_local, data_only=False)
                        sheet_match = re.match(r"'([^']+)'!(.+)", answer_position)
                        if sheet_match:
                            sheet_name = sheet_match.group(1)
                            cell_range = sheet_match.group(2)
                            ws_input = wb_input[sheet_name]
                        else:
                            cell_range = answer_position
                            ws_input = wb_input.active
                        
                        # Collect input answer values
                        input_coords_lines = []
                        input_has_formulas = False
                        for row in ws_input[cell_range]:
                            for cell in row:
                                if hasattr(cell, 'data_type') and cell.data_type == 'f':
                                    input_coords_lines.append(f"{cell.coordinate}: FORMULA = {cell.value}")
                                    input_has_formulas = True
                                else:
                                    input_coords_lines.append(f"{cell.coordinate}: {cell.value}")
                                input_answer_values_raw.append(cell.value)
                        
                        input_answer_content = '\n'.join(input_coords_lines)
                        input_non_empty = [v for v in input_answer_values_raw if v not in (None, "")]
                        input_unique_vals = set(input_non_empty)
                        input_numeric_vals = [float(v) for v in input_non_empty if isinstance(v, (int,float))]
                        input_summary_json = {
                            "total_cells": len(input_answer_values_raw),
                            "non_empty_count": len(input_non_empty),
                            "unique_count": len(input_unique_vals),
                            "all_same": len(input_unique_vals) == 1,
                            "has_formulas": input_has_formulas,
                            "sample_values": list(input_non_empty[:10]),
                            "numeric_min": min(input_numeric_vals) if input_numeric_vals else None,
                            "numeric_max": max(input_numeric_vals) if input_numeric_vals else None,
                        }
                        wb_input.close()
                except Exception as e:
                    input_answer_content = f"Error reading input: {str(e)}"
                
                # SECOND: Read output file's answer column
                try:
                    if os.path.exists(output_local):
                        wb = openpyxl.load_workbook(output_local, data_only=False)
                        sheet_match = re.match(r"'([^']+)'!(.+)", answer_position)
                        if sheet_match:
                            sheet_name = sheet_match.group(1)
                            cell_range = sheet_match.group(2)
                            ws = wb[sheet_name]
                        else:
                            cell_range = answer_position
                            ws = wb.active
                        # Extract raw cells
                        def _col_letter_to_index(col_letters: str) -> int:
                            from openpyxl.utils import column_index_from_string
                            return column_index_from_string(col_letters)
                        def _parse_range(r: str):
                            if ':' not in r:
                                return r, r
                            return r.split(':', 1)
                        start_ref, end_ref = _parse_range(cell_range)
                        # Derive column letters
                        import string
                        import itertools
                        def _split_ref(ref: str):
                            m = re.match(r"([A-Z]+)([0-9]+)", ref)
                            return m.group(1), int(m.group(2)) if m else (None, None)
                        start_col_letters, start_row_num = _split_ref(start_ref)
                        end_col_letters, end_row_num = _split_ref(end_ref)
                        start_col_idx = _col_letter_to_index(start_col_letters)
                        end_col_idx = _col_letter_to_index(end_col_letters)
                        # Collect values
                        for row in ws[cell_range]:
                            for cell in row:
                                answer_values_raw.append(cell.value)
                        # Build formatted content listing coordinate:value (detect formulas)
                        coords_lines = []
                        has_formulas = False
                        for row in ws[cell_range]:
                            for cell in row:
                                if hasattr(cell, 'data_type') and cell.data_type == 'f':
                                    coords_lines.append(f"{cell.coordinate}: FORMULA = {cell.value}")
                                    has_formulas = True
                                else:
                                    coords_lines.append(f"{cell.coordinate}: {cell.value}")
                        answer_content = '\n'.join(coords_lines)
                        non_empty = [v for v in answer_values_raw if v not in (None, "")]
                        unique_vals = set(non_empty)
                        numeric_vals = [float(v) for v in non_empty if isinstance(v, (int,float))]
                        summary_json = {
                            "total_cells": len(answer_values_raw),
                            "non_empty_count": len(non_empty),
                            "unique_count": len(unique_vals),
                            "all_same": len(unique_vals) == 1,
                            "has_formulas": has_formulas,
                            "sample_values": list(non_empty[:10]),
                            "numeric_min": min(numeric_vals) if numeric_vals else None,
                            "numeric_max": max(numeric_vals) if numeric_vals else None,
                            "numeric_mean": (sum(numeric_vals)/len(numeric_vals)) if numeric_vals else None,
                        }
                        # Neighbor column leak detection (only if single column range)
                        if start_col_idx == end_col_idx:
                            right_col_idx = end_col_idx + 1
                            # Only check if within sheet bounds
                            if right_col_idx <= ws.max_column:
                                from openpyxl.utils import get_column_letter
                                right_letter = get_column_letter(right_col_idx)
                                leak_values = []
                                for r in range(start_row_num, end_row_num + 1):
                                    cv = ws.cell(row=r, column=right_col_idx).value
                                    if cv not in (None, ""):
                                        leak_values.append(cv)
                                if leak_values:
                                    neighbor_alert = {
                                        "right_column": right_letter,
                                        "non_empty_count": len(leak_values),
                                        "sample": leak_values[:10]
                                    }
                        wb.close()
                except Exception as e:
                    answer_content = f"Error reading output: {str(e)}"
                import json as _json
                answer_summary_block = _json.dumps(summary_json, ensure_ascii=False, indent=2)
                input_summary_block = _json.dumps(input_summary_json, ensure_ascii=False, indent=2)
                neighbor_alert_block = _json.dumps(neighbor_alert, ensure_ascii=False, indent=2) if neighbor_alert else "None"
                validation_prompt = f"""You are SheetCopilot v2 in CODE VALIDATION stage.

The code executed SUCCESSFULLY. Evaluate the semantic correctness of results.

📋 **ORIGINAL TASK**: {instruction}

📊 **OBSERVED INPUT (truncated, CHECK FORMAT REFERENCE)**:
{observation['response'][:600]}

📋 **IMPLEMENTATION PLAN (truncated)**:
{plan['response'][:600]}

💻 **EXECUTED CODE**:
```python
{implementation['code']}
```

✅ **RAW EXECUTION STDOUT/LOG OUTPUT**:
```
{exec_result}
```

🎯 **INPUT ANSWER COLUMN PATTERN (GROUND TRUTH REFERENCE in {answer_position})**:
```
{input_answer_content[:800]}
```

📊 **INPUT ANSWER SUMMARY**:
```json
{input_summary_block}
```

📌 **OUTPUT RESULT CELLS (Generated by your code in {answer_position})**:
```
{answer_content[:800]}
```

📊 **OUTPUT RESULT SUMMARY**:
```json
{answer_summary_block}
```

🛑 **NEIGHBOR COLUMN LEAK CHECK (right column)**:
```json
{neighbor_alert_block}
```

**EVALUATION INSTRUCTIONS (CRITICAL - FOLLOW EXACTLY)**:

1. **DISCOVER INPUT PATTERN (HIGHEST PRIORITY)**:
   - Analyze the "INPUT ANSWER COLUMN PATTERN" section carefully
   - What format does it show? Formulas (=IF, =VLOOKUP, etc.) or plain values?
   - If formulas: What formula pattern? (e.g., all rows use =IF with same structure)
   - If plain values: What DATA TYPE? (numeric, text, dates, percentages)
   - Is there case sensitivity? (UPPER/lower/Mixed)
   - Are there delimiters? (comma, space, etc.)
   
2. **DATA TYPE VALIDATION (CRITICAL)**:
   - Check INPUT ANSWER SUMMARY → "has_formulas" field
   - If has_formulas=true → INPUT cells contain FORMULA STRINGS (start with =)
   - If has_formulas=false → INPUT cells contain PLAIN VALUES
   - **KEY RULE**: OUTPUT must match INPUT's data type (formula vs value)
   - Look at "sample_values" to see actual content type
   
3. **COMPARE WITH USER INTENT**:
   - Does the task instruction ask for formulas or values?
   - Does the input pattern match what user expects?
   - **KEY INSIGHT**: Input pattern shows the DESIRED OUTPUT FORMAT!
   - If observation shows "FORMULA PATTERN DETECTED" → verify INPUT has formulas
   
4. **VALIDATE OUTPUT AGAINST INPUT PATTERN**:
   - Compare OUTPUT RESULT SUMMARY with INPUT ANSWER SUMMARY:
     * has_formulas field must match
     * If INPUT has_formulas=true but OUTPUT has_formulas=false → CRITICAL ERROR
     * Data type patterns must align (str→str, int→int, float→float)
   - If INPUT shows specific text casing → OUTPUT must match casing
   - If INPUT shows numeric pattern → OUTPUT must match range/pattern
   
5. **DETECT COMMON DATA TYPE ERRORS**:
   - ❌ INPUT has_formulas=true but OUTPUT has_formulas=false → FAILED 
      (code EVALUATED formulas and wrote VALUES instead of FORMULA STRINGS)
   - ❌ INPUT has_formulas=false but OUTPUT has_formulas=true → FAILED
      (code wrote formulas when plain values expected)
   - ❌ INPUT sample_values shows strings but OUTPUT shows numbers → FAILED (wrong type conversion)
   - ❌ Input has UPPER case but output has lower → FAILED (wrong case transformation)
   - ❌ Neighbor column leak detected → FAILED (code wrote to wrong columns)

**RESPOND WITH ONE OF:**

Option A (PASSED - output pattern matches input pattern):
```
VALIDATION PASSED
Pattern Analysis:
- Input shows: [describe input pattern briefly]
- Output shows: [describe output pattern briefly]
- Match confirmed: [why they align]
```

Option B (FAILED - pattern mismatch or errors):
```
VALIDATION FAILED
Pattern Mismatch:
- Input pattern: [what input shows]
- Output pattern: [what output shows]
- Root cause: [why code generated wrong format]

Issues:
1. [specific issue]
2. [...]

Fix Strategy:
- [how to match input pattern]

CORRECTED CODE:
```python
[provide fully revised code that matches input pattern]
```
```

Return ONLY one option.
"""
        
        except Exception as e:
            exec_result = f"Exception during validation execution: {str(e)}"
            validation_prompt = f"""You are SheetCopilot v2 in CODE VALIDATION stage.

Code execution raised an EXCEPTION. Please fix the code.

💻 **CODE** (has issues):
```python
{implementation['code']}
```

❌ **EXCEPTION**:
```
{exec_result}
```

Provide CORRECTED code:
"""
        
        # Generate validation response (logging disabled)
        messages = [
            observation['prompt'], observation['response'],
            plan['prompt'], plan['response'],
            implementation['prompt'], implementation['response'],
            validation_prompt
        ]
        response = get_llm_response(messages, self.opt)
        
        # Check if validation passed or code was corrected
        validation_passed = "VALIDATION PASSED" in response.upper()
        corrected_code = extract_code(response) if not validation_passed else None
        
        if self.enable_timing:
            stage_time = time.time() - stage_start
            self.stage_timings['stage_5_validation'] = stage_time
            if self.debug:
                print(f"[DEBUG] ✅ Stage 5 completed in {stage_time:.2f}s")
            self.log_stage("5️⃣ CODE VALIDATION", 
                          "Execute and verify results reasonableness", 
                          stage_time)
        
        return {
            'prompt': validation_prompt,
            'response': response,
            'passed': validation_passed,
            'corrected_code': corrected_code,
            'execution_result': exec_result,
            'issues_found': self._extract_issues(response),
            'answer_values_summary': summary_json if 'summary_json' in locals() else {},
            'neighbor_alert': neighbor_alert if 'neighbor_alert' in locals() else None
        }
    
    def stage_6_execution_and_revision(self, validation: Dict[str, Any],
                                      implementation: Dict[str, Any],
                                      observation: Dict[str, Any],
                                      plan: Dict[str, Any],
                                      file_path: str, output_path: str,
                                      instruction: str) -> Dict[str, Any]:
        """
        Stage 6: Execution and Revision - Execute code with retry mechanism
        
        Execution with smart revision:
        - Try validated/corrected code first
        - If errors, analyze and fix up to max_revisions times
        - Learn from execution feedback
        - Adapt to runtime errors
        """
        import time
        stage_start = time.time() if self.enable_timing else None
        
        if self.debug:
            print(f"\n[DEBUG] 🚀 Starting Stage 6: Final Execution & Revision")
        
        # Print stage header at the beginning
        if self.enable_timing:
            self.log_stage("6️⃣ EXECUTION & REVISION", 
                          "Running code with intelligent error recovery", 
                          None)  # Time will be updated at the end
        
        # Use corrected code from validation if available (handle None case)
        if validation is not None:
            code_to_execute = validation.get('corrected_code') or implementation['code']
        else:
            code_to_execute = implementation['code']
        
        # Replace hardcoded paths before execution (important for code reuse)
        code_to_execute = self._replace_hardcoded_paths(code_to_execute, file_path, output_path)
        
        for revision_num in range(self.max_revisions + 1):
            try:
                result = exec_code(self.client, code_to_execute)
                
                # Check for errors in output
                has_error = 'Error' in result or 'Traceback' in result or '❌' in result
                
                # Only log errors or final success
                if has_error:
                    self.logger.info(f"[EXECUTION ERROR - Attempt {revision_num + 1}]\n{result}")
                
                if not has_error:
                    if self.debug and self.enable_timing:
                        elapsed = time.time() - stage_start
                        print(f"[DEBUG] ✅ Stage 6 execution successful (attempt {revision_num + 1}, {elapsed:.2f}s)")
                    # Calculate formulas if output file exists (Windows only)
                    output_local = output_path.replace('/mnt/data/', '../data/')
                    if os.path.exists(output_local):
                        calculate_formulas(output_local, self.logger)
                        if getattr(self.opt, 'excel_recalc', False):
                            recalc_workbook(
                                input_path=output_local,
                                output_path=output_local,
                                materialize_dynamic=getattr(self.opt, 'materialize_dynamic', False),
                                strip_formula=getattr(self.opt, 'strip_dynamic_formula', False),
                                logger=self.logger,
                            )
                    if self.enable_timing:
                        stage_time = time.time() - stage_start
                        self.stage_timings['stage_6_execution'] = stage_time
                        if self.debug:
                            print(f"[DEBUG] ✅ Stage 6 completed in {stage_time:.2f}s (success after {revision_num} revisions)")
                    return {
                        'success': True,
                        'result': result,
                        'final_code': code_to_execute,
                        'revisions_needed': revision_num
                    }
                
                # Error occurred, need revision
                if revision_num < self.max_revisions:
                    if self.debug:
                        print(f"[DEBUG] ⚠️  Stage 6 error detected, revising code (attempt {revision_num + 1}/{self.max_revisions})")
                    code_to_execute = self._revise_code(
                        code_to_execute, result, observation, plan, instruction, 
                        file_path, output_path
                    )
                else:
                    self.logger.info(f"❌ Max revisions reached")
                    if self.enable_timing:
                        stage_time = time.time() - stage_start
                        self.stage_timings['stage_6_execution'] = stage_time
                    return {
                        'success': False,
                        'result': result,
                        'final_code': code_to_execute,
                        'revisions_needed': revision_num
                    }
                    
            except Exception as e:
                error_msg = f"Exception during execution: {str(e)}"
                self.logger.info(f"[EXECUTION EXCEPTION]\n{error_msg}")
                
                if revision_num < self.max_revisions:
                    code_to_execute = self._revise_code(
                        code_to_execute, error_msg, observation, plan, instruction,
                        file_path, output_path
                    )
                else:
                    if self.enable_timing:
                        stage_time = time.time() - stage_start
                        self.stage_timings['stage_6_execution'] = stage_time
                    return {
                        'success': False,
                        'result': error_msg,
                        'final_code': code_to_execute,
                        'revisions_needed': revision_num
                    }
        
        if self.enable_timing:
            stage_time = time.time() - stage_start
            self.stage_timings['stage_6_execution'] = stage_time
        return {
            'success': False,
            'result': 'Unexpected execution flow',
            'final_code': code_to_execute,
            'revisions_needed': self.max_revisions
        }
    
    def _revise_code(self, current_code: str, error_output: str,
                    observation: Dict[str, Any], plan: Dict[str, Any],
                    instruction: str, file_path: str, output_path: str) -> str:
        """Internal method to revise code based on execution errors"""
        
        revision_prompt = f"""You are SheetCopilot v2 in ERROR RECOVERY mode.


🎯 **TASK**: {instruction}

📊 **SPREADSHEET STRUCTURE** (observed facts):
{observation['result'][:600]}

📋 **ORIGINAL PLAN**:
{plan['response'][:600]}

💻 **CURRENT CODE** (has errors):
```python
{current_code}
```

❌ **EXECUTION ERROR**:
{error_output}

**YOUR DEBUGGING TASK**:
1. Carefully read the error traceback
2. Identify root cause (common issues in real-world spreadsheets):
   - Wrong cell reference (maybe assumed A1 instead of actual position)
   - Sheet name mismatch
   - Index out of range (table smaller than expected)
   - AttributeError (cell is None/empty)
   - TypeError (wrong data type, need int() or float())
   - KeyError (sheet doesn't exist)
   - Excel formula syntax errors:
     * @ prefix before function names (e.g., @XLOOKUP should be XLOOKUP)
     * @ prefix before sheet names (e.g., @Sheet1 should be Sheet1)
     * Wrong string concatenation (e.g., "*&A1&*" should be "*"&A1&"*")
     * Missing quotes around string literals in formulas

3. Fix the code COMPLETELY
4. Ensure fix addresses the root cause, not just symptoms

**CRITICAL REMINDERS**:
- Use OBSERVED cell positions, not assumptions
- Check cell.value is not None before operations
- Validate indices are within actual range
- Use correct sheet names from observation
- ⚠️ AVOID CIRCULAR REFERENCES: Do NOT reference target cells in formulas
- ⚠️ EXCEL FORMULA SYNTAX: NO @ prefix, correct string concatenation with &

**Generate FIXED code**:
"""
        
        # Generate revised code (prompt logging disabled)
        messages = [
            observation['prompt'], observation['response'],
            plan['prompt'], plan['response'],
            revision_prompt
        ]
        response = get_llm_response(messages, self.opt)
        
        revised_code = extract_code(response)
        # Replace with current test case paths (in case LLM hardcoded old paths)
        revised_code = self._replace_hardcoded_paths(revised_code, file_path, output_path)
        # Log revised code (KEY INFO for debugging)
        self.logger.info(f"[REVISED CODE]\n{revised_code}")
        
        return revised_code
    
    # Helper methods for parsing LLM responses
    def _parse_requirements(self, response: str) -> Dict[str, str]:
        """Extract structured requirements from understanding response"""
        sections = {}
        current_section = None
        current_content = []
        
        for line in response.split('\n'):
            if line.startswith('##'):
                if current_section:
                    sections[current_section] = '\n'.join(current_content).strip()
                current_section = line.strip('# ').strip()
                current_content = []
            elif current_section:
                current_content.append(line)
        
        if current_section:
            sections[current_section] = '\n'.join(current_content).strip()
        
        return sections
    
    def _extract_steps(self, response: str) -> List[str]:
        """Extract step-by-step plan from planning response"""
        steps = []
        for line in response.split('\n'):
            if line.strip().startswith('###'):
                steps.append(line.strip('# ').strip())
        return steps
    
    def _extract_issues(self, response: str) -> List[str]:
        """Extract validation issues from validation response"""
        issues = []
        for line in response.split('\n'):
            if '✗' in line or '[✗]' in line or '- [ ]' in line:
                issues.append(line.strip())
        return issues
    
    def _replace_hardcoded_paths(self, code: str, file_path: str, output_path: str) -> str:
        """Replace hardcoded file paths in generated code with current test case paths"""
        import re
        
        # Pattern 1: openpyxl.load_workbook('...')
        code = re.sub(
            r"openpyxl\.load_workbook\(['\"]([^'\"]+)['\"]\)",
            f"openpyxl.load_workbook('{file_path}')",
            code
        )
        
        # Pattern 2: wb.save('...')
        code = re.sub(
            r"wb\.save\(['\"]([^'\"]+)['\"]\)",
            f"wb.save('{output_path}')",
            code
        )
        
        # Pattern 3: Variable assignment like input_path = '...'
        code = re.sub(
            r"input_path\s*=\s*['\"]([^'\"]+)['\"]",
            f"input_path = '{file_path}'",
            code
        )
        code = re.sub(
            r"output_path\s*=\s*['\"]([^'\"]+)['\"]",
            f"output_path = '{output_path}'",
            code
        )
        
        return code
    
    def solve_task(self, task_data: Dict[str, Any], dataset_path: str) -> Dict[str, Any]:
        """
        Main pipeline: Process one task through all stages
        
        Enhanced 6-stage pipeline specifically for SpreadsheetBench:
        1. Deep Observation - Understand non-standard structure
        2. Instruction Understanding - Parse complex natural language
        3. Solution Planning - Design robust approach
        4. Code Implementation - Generate Python code
        5. Code Validation - Static analysis
        6. Execution & Revision - Run with retry mechanism
        """
        task_id = task_data['id']
        self.logger.info(f"\n{'='*60}\n🚀 Task {task_id}\n{'='*60}")
        
        if self.debug:
            print(f"\n{'='*80}")
            print(f"[DEBUG] 🎯 Starting Task {task_id}: {task_data['instruction'][:80]}...")
            print(f"{'='*80}")
        
        import time
        task_start = time.time() if self.enable_timing else None

        all_case_results = []
        aggregate_conversation = []
        total_revisions = 0
        overall_success = True
        final_code_last = None
        
        # Reset stage timings for this task
        self.stage_timings = {}
        
        # Shared code generation results (only generate once for test case 1)
        shared_understanding = None
        shared_plan = None
        shared_implementation = None
        shared_validation = None

        for test_case_idx in range(1, 4):
            file_name = f"{test_case_idx}_{task_data['spreadsheet_path'].lstrip('spreadsheet/')}_input.xlsx"
            input_path = f"/mnt/data/{self.opt.dataset}/{task_data['spreadsheet_path']}/{file_name}"
            output_path = f"/mnt/data/{self.opt.dataset}/outputs/sheetcopilot_{self.opt.model}/{file_name.replace('_input.xlsx', '_output.xlsx')}"
            if self.debug:
                print(f"\n[DEBUG] 📝 Processing test case {test_case_idx}/3")
            # Processing test case (logging reduced)

            conversation = []
            # Reset stage history per test case (must be list for append)
            self.stage_history = []
            try:
                # Stage 1: Deep Observation (always run - each test case has different input file)
                observation = self.stage_1_deep_observation(
                    input_path,
                    task_data['instruction'],
                    task_data['answer_position'],
                    task_data['instruction_type']
                )
                # Only add prompt summary to conversation, not full response (too verbose)
                conversation.append(f"[Stage 1 Observation completed for {input_path}]")
                if not observation['success']:
                    self.logger.info(f"⚠️ Observation failed: test {test_case_idx}")
                    all_case_results.append({
                        'test_case_index': test_case_idx,
                        'success': False,
                        'revisions_needed': 0,
                        'final_code': ''
                    })
                    overall_success = False
                    aggregate_conversation.extend(conversation)
                    continue

                # Stage 2-5: Only generate for test case 1, reuse for 2 and 3
                if test_case_idx == 1:
                    # Generate code with LLM for test case 1
                    
                    try:
                        # Stage 2: Instruction Understanding
                        shared_understanding = self.stage_2_instruction_understanding(
                            observation,
                            task_data['instruction'],
                            task_data['instruction_type']
                        )
                        conversation.append("[Stage 2 Understanding completed]")

                        # Stage 3: Solution Planning
                        shared_plan = self.stage_3_solution_planning(
                            observation, shared_understanding,
                            input_path, output_path,
                            task_data['answer_position']
                        )
                        conversation.append("[Stage 3 Planning completed]")

                        # Stage 4: Code Implementation
                        shared_implementation = self.stage_4_code_implementation(
                            observation, shared_understanding, shared_plan,
                            input_path, output_path,
                            task_data['answer_position']
                        )
                        conversation.append(f"[Stage 4 Implementation completed - Code length: {len(shared_implementation['code'])} chars]")

                        # Stage 5: Code Validation (with execution and result verification)
                        shared_validation = self.stage_5_code_validation(
                            shared_implementation, observation, shared_plan,
                            input_path, output_path,
                            task_data['instruction'], task_data['answer_position']
                        )
                        conversation.append(f"[Stage 5 Validation: {'PASSED' if shared_validation['passed'] else 'ISSUES FOUND'}]")
                    
                    except Exception as llm_error:
                        self.logger.error(f"❌ LLM stages (2-5) failed for test case 1: {str(llm_error)}")
                        # Mark all test cases as failed since we can't generate code
                        for tc_idx in range(1, 4):
                            all_case_results.append({
                                'test_case_index': tc_idx,
                                'success': False,
                                'revisions_needed': 0,
                                'final_code': ''
                            })
                        overall_success = False
                        aggregate_conversation.extend(conversation)
                        break  # Exit the test case loop
                else:
                    # Reuse code from test case 1
                    conversation.append(f"[Reusing code generated for test case 1]")

                # Stage 6: Execution & Revision (always run - different input/output paths)
                execution = self.stage_6_execution_and_revision(
                    shared_validation, shared_implementation, observation, shared_plan,
                    input_path, output_path, task_data['instruction']
                )
                conversation.append(execution['result'])
                self.logger.info(f"Test {test_case_idx}: {'✅' if execution['success'] else '❌'} (revisions: {execution['revisions_needed']})")
                final_code_last = execution['final_code']
                total_revisions += execution['revisions_needed']
                
                # Update shared code with revised version for next test case
                if execution['success'] and execution['final_code']:
                    # Store the successful code, but remove test-case-specific paths
                    # Next test case will have paths replaced via _replace_hardcoded_paths
                    shared_implementation['code'] = execution['final_code']
                    if shared_validation is not None and shared_validation.get('corrected_code'):
                        shared_validation['corrected_code'] = execution['final_code']
                if not execution['success']:
                    overall_success = False
                all_case_results.append({
                    'test_case_index': test_case_idx,
                    'success': execution['success'],
                    'revisions_needed': execution['revisions_needed'],
                    'final_code': execution['final_code'] or ''
                })
            except Exception as e:
                self.logger.info(f"❌ Exception test {test_case_idx}: {str(e)}")
                import traceback
                traceback.print_exc()
                overall_success = False
                all_case_results.append({
                    'test_case_index': test_case_idx,
                    'success': False,
                    'revisions_needed': 0,
                    'final_code': ''
                })
            aggregate_conversation.extend(conversation)

        if self.enable_timing:
            task_time = time.time() - task_start
            self.logger.info(f"Task time: {task_time:.2f}s")
            if self.debug:
                print(f"\n[DEBUG] ⏱️  Task {task_id} total time: {task_time:.2f}s")
                print(f"[DEBUG] Stage timings:")
                for stage_name, stage_time in self.stage_timings.items():
                    print(f"[DEBUG]   - {stage_name}: {stage_time:.2f}s")
        
        self.logger.info(f"Task {task_id}: {'✅ SUCCESS' if overall_success else '❌ FAILED'}")
        if self.debug:
            print(f"[DEBUG] {'✅ Task SUCCEEDED' if overall_success else '❌ Task FAILED'}\n{'='*60}")
        # Return combined result; final_code uses last successful code if any
        combined_result = self._create_result(task_data, aggregate_conversation, final_code_last, overall_success, total_revisions)
        combined_result['per_test_case'] = all_case_results
        if self.enable_timing:
            combined_result['timing'] = {
                'total_time': task_time,
                'stage_timings': self.stage_timings.copy()
            }
        return combined_result
    
    def _create_result(self, task_data: Dict[str, Any], conversation: List[str],
                      final_code: Optional[str], success: bool, 
                      revisions: int) -> Dict[str, Any]:
        """Create standardized result dictionary"""
        return {
            'id': task_data['id'],
            'instruction_type': task_data['instruction_type'],
            'conversation': conversation,
            'solution': final_code or '',
            'success': success,
            'revisions_needed': revisions,
            'stage_history': self.stage_history.copy()
        }


def parse_option():
    """Parse command line arguments"""
    parser = argparse.ArgumentParser("SheetCopilot v2 - Enhanced for Real-World Spreadsheets")
    
    parser.add_argument('--model', type=str, required=True, help='LLM model name')
    parser.add_argument('--api_key', type=str, default="", help='API key for model')
    parser.add_argument('--base_url', type=str, default="", help='Base URL for model API')
    parser.add_argument('--dataset', type=str, default="test1", help='Dataset name')
    parser.add_argument('--code_exec_url', type=str, 
                       default="http://localhost:8080/execute", 
                       help='Code execution Docker URL')
    parser.add_argument('--conv_id', type=str, default="EVAL", 
                       help='Conversation ID for code execution')
    parser.add_argument('--max_revisions', type=int, default=3,
                       help='Maximum number of code revision attempts')
    parser.add_argument('--log_dir', type=str, default='../log',
                       help='Directory for log files')
    parser.add_argument('--enable_timing', action='store_true', default=True,
                       help='Enable stage timing for performance analysis')
    parser.add_argument('--disable_timing', dest='enable_timing', action='store_false',
                       help='Disable stage timing')
    parser.add_argument('--debug', action='store_true', default=True,
                       help='Enable debug output showing stage progress and timing in terminal')
    parser.add_argument('--excel-recalc', action='store_true', default=False,
                       help='After Docker execution, open workbook in real Excel to force full recalc (dynamic arrays spill)')
    parser.add_argument('--materialize-dynamic', action='store_true', default=False,
                       help='With --excel-recalc: convert dynamic spill ranges to static values for openpyxl compatibility')
    parser.add_argument('--strip-dynamic-formula', action='store_true', default=False,
                       help='With --materialize-dynamic: replace source spill formula cell with its calculated value')
    
    return parser.parse_args()


def main():
    """Main execution function"""
    opt = parse_option()
    print(f"SheetCopilot v2 Configuration:\n{opt}\n")
    
    # Setup logging
    logger = setup_logger(opt.log_dir, opt.model, opt.dataset)
    logger.info(f"Starting SheetCopilot v2 with config: {opt}")
    
    # Load dataset
    dataset_path = os.path.abspath(f'../data/{opt.dataset}')
    with open(f'{dataset_path}/dataset.json', 'r') as fp:
        dataset = json.load(fp)
    
    logger.info(f"Loaded {len(dataset)} tasks from {opt.dataset}")
    
    # Prepare output directories
    output_dir = f'{dataset_path}/outputs/sheetcopilot_{opt.model}'
    os.makedirs(output_dir, exist_ok=True)
    os.chmod(output_dir, 0o777)
    
    conv_file = f'{dataset_path}/outputs/conv_sheetcopilot_{opt.model}.jsonl'
    
    # Initialize copilot
    copilot = SheetCopilotV2(opt, logger)
    
    # Process tasks
    success_count = 0
    total_revisions = 0
    
    for task_data in tqdm(dataset, desc="Processing tasks"):
        result = copilot.solve_task(task_data, dataset_path)
        
        # Save conversation
        with open(conv_file, 'a+', encoding='utf-8') as fp:
            fp.write(json.dumps(result, ensure_ascii=False) + '\n')
        
        if result['success']:
            success_count += 1
        total_revisions += result['revisions_needed']
    
    # Summary
    success_rate = success_count / len(dataset) * 100
    avg_revisions = total_revisions / len(dataset)
    
    logger.info(f"\n{'='*100}")
    logger.info(f"FINAL RESULTS:")
    logger.info(f"Total tasks: {len(dataset)}")
    logger.info(f"Successful: {success_count}/{len(dataset)} ({success_rate:.1f}%)")
    logger.info(f"Average revisions: {avg_revisions:.2f}")
    logger.info(f"='*100")
    
    print(f"\n✅ SheetCopilot v2 completed!")
    print(f"Success rate: {success_rate:.1f}%")
    print(f"Results saved to: {conv_file}")


if __name__ == '__main__':
    main()
