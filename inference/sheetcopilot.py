"""
SheetCopilot: A multi-stage reasoning system for spreadsheet manipulation
Implements Observing → Proposing → Revising → Executing cycle with tool usage
"""

import os
import json
import logging
import argparse
import pandas as pd
from tqdm import tqdm
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple

from llm_api import get_llm_response
from code_exec import get_exec_client, extract_code, exec_code


# Configure logging
def setup_logger(log_dir: str, model_name: str) -> logging.Logger:
    """Setup detailed logging for debugging and optimization"""
    os.makedirs(log_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = f"{log_dir}/sheetcopilot_{model_name}_{timestamp}.log"
    
    logger = logging.getLogger('SheetCopilot')
    logger.setLevel(logging.DEBUG)
    
    # File handler with detailed format
    fh = logging.FileHandler(log_file, encoding='utf-8')
    fh.setLevel(logging.DEBUG)
    
    # Console handler with simpler format
    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    
    # Detailed formatter
    formatter = logging.Formatter(
        '[%(asctime)s] [%(name)s] [%(levelname)s] [%(funcName)s:%(lineno)d]\n%(message)s\n',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    fh.setFormatter(formatter)
    ch.setFormatter(logging.Formatter('[%(levelname)s] %(message)s'))
    
    logger.addHandler(fh)
    logger.addHandler(ch)
    
    return logger


class SheetCopilot:
    """Main SheetCopilot system with multi-stage reasoning"""
    
    def __init__(self, opt, logger: logging.Logger):
        self.opt = opt
        self.logger = logger
        self.client = get_exec_client(opt.code_exec_url, opt.conv_id)
        
        # Stage tracking
        self.current_stage = None
        self.stage_history = []
        
    def log_stage(self, stage_name: str, content: str):
        """Log stage information with structured format"""
        self.logger.info(f"\n{'='*80}\n[STAGE] {stage_name}\n{'='*80}")
        self.logger.info(content)
        self.stage_history.append({
            'stage': stage_name,
            'content': content,
            'timestamp': datetime.now().isoformat()
        })
    
    def stage_1_observing(self, file_path: str, instruction: str, 
                          answer_position: str, instruction_type: str) -> Dict[str, Any]:
        """
        Stage 1: Observing - Let LLM understand the spreadsheet state
        Returns: observation_result with spreadsheet state information
        """
        self.current_stage = "OBSERVING"
        self.log_stage("OBSERVING", f"File: {file_path}\nTask: {instruction}")
        
        # Build observation prompt with inline tool implementations
        observation_prompt = f"""You are SheetCopilot, an expert spreadsheet assistant. You are in the OBSERVING stage.

**Task**: {instruction}
**Instruction Type**: {instruction_type}
**Target Position**: {answer_position}
**File Path**: {file_path}

Your goal in this stage is to understand the spreadsheet state by writing Python code using openpyxl library.

**Available Operations** (use openpyxl directly):
1. Load workbook and get sheet names
2. Get sheet dimensions (max_row, max_column)
3. Read cell ranges and values
4. Check cell formats (font, fill, number_format)
5. Search for specific values
6. Get column/row data

**Example Code Pattern**:
```python
import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('{file_path}')

# Get sheet names
print("Sheet names:", wb.sheetnames)

# Work with active sheet
ws = wb.active
print(f"Dimensions: {{ws.max_row}} rows x {{ws.max_column}} columns")

# Read target range (adjust based on answer_position)
print(f"Target range '{answer_position}':")
for row in ws['{answer_position}']:
    values = [cell.value for cell in row]
    print(values)

# Read headers or relevant data
# ... add more observations as needed

wb.close()
```

**Instructions**:
1. Write Python code to observe the spreadsheet structure
2. Focus on the target position: {answer_position}
3. Check what data exists and what needs to be filled/modified
4. Use print() to show your observations

Generate complete Python code for observation:
"""
        
        self.logger.debug(f"[OBSERVING PROMPT]\n{observation_prompt}")
        
        # Get LLM response
        response = get_llm_response([observation_prompt], self.opt)
        self.logger.debug(f"[OBSERVING RESPONSE]\n{response}")
        
        # Execute observation code
        code = extract_code(response)
        self.logger.debug(f"[OBSERVING CODE]\n{code}")
        
        try:
            observation_result = exec_code(self.client, code)
            self.logger.info(f"[OBSERVING RESULT]\n{observation_result}")
        except Exception as e:
            self.logger.error(f"[OBSERVING ERROR] {str(e)}")
            observation_result = f"Error during observation: {str(e)}"
        
        return {
            'prompt': observation_prompt,
            'response': response,
            'code': code,
            'result': observation_result,
            'success': 'Error' not in observation_result
        }
    
    def stage_2_proposing(self, observation: Dict[str, Any], file_path: str, 
                         output_path: str, instruction: str, 
                         answer_position: str, instruction_type: str) -> Dict[str, Any]:
        """
        Stage 2: Proposing - LLM proposes atomic actions based on observation
        Returns: proposed actions and implementation code
        """
        self.current_stage = "PROPOSING"
        self.log_stage("PROPOSING", f"Based on observation, propose solution for: {instruction}")
        
        # Build proposing prompt
        proposing_prompt = f"""You are SheetCopilot in the PROPOSING stage.

**Original Task**: {instruction}
**Instruction Type**: {instruction_type}
**Target Position**: {answer_position}
**Input File**: {file_path}
**Output File**: {output_path}

**Observation Results**:
{observation['result']}

Based on your observation, you now need to:
1. Break down the task into atomic actions (e.g., read data, calculate, write result, format cells)
2. Propose a clear step-by-step plan
3. Generate COMPLETE Python code to implement the solution

**Requirements**:
- Use openpyxl library for all spreadsheet operations
- Only modify cells within the target position: {answer_position}
- For Cell-Level tasks: modify specific cells
- For Sheet-Level tasks: may modify entire worksheet within range
- Load from {file_path} and save result to: {output_path}
- Include all necessary imports (openpyxl, pandas, numpy, etc.)
- Ensure code is complete and can run independently

**Code Template**:
```python
import openpyxl
# Add other imports as needed

# Load input file
wb = openpyxl.load_workbook('{file_path}')
ws = wb.active  # or specify sheet name

# Your solution code here
# 1. Read necessary data
# 2. Perform calculations/operations
# 3. Write results to target cells

# Save output file
wb.save('{output_path}')
wb.close()
print("Successfully saved to {output_path}")
```

**Output Format**:
First, describe your plan:
## Plan
1. Action 1: ...
2. Action 2: ...

Then provide COMPLETE implementation code:
```python
# Your complete implementation here
```

Generate your proposal now:
"""
        
        self.logger.debug(f"[PROPOSING PROMPT]\n{proposing_prompt}")
        
        # Get LLM response
        messages = [observation['prompt'], observation['response'], proposing_prompt]
        response = get_llm_response(messages, self.opt)
        self.logger.debug(f"[PROPOSING RESPONSE]\n{response}")
        
        # Extract code
        code = extract_code(response)
        self.logger.debug(f"[PROPOSING CODE]\n{code}")
        
        return {
            'prompt': proposing_prompt,
            'response': response,
            'code': code,
            'plan_extracted': self._extract_plan(response)
        }
    
    def stage_3_revising(self, observation: Dict[str, Any], 
                        proposal: Dict[str, Any], 
                        execution_result: str,
                        file_path: str, output_path: str,
                        instruction: str, answer_position: str) -> Dict[str, Any]:
        """
        Stage 3: Revising - Revise the proposed action based on execution feedback
        Returns: revised code or validation result
        """
        self.current_stage = "REVISING"
        self.log_stage("REVISING", "Analyzing execution result and revising if needed")
        
        # Check if revision is needed
        needs_revision = 'Error' in execution_result or 'Traceback' in execution_result
        
        if not needs_revision:
            self.logger.info("[REVISING] No errors detected, validation passed")
            return {
                'needs_revision': False,
                'execution_result': execution_result,
                'validation': 'SUCCESS'
            }
        
        # Build revising prompt
        revising_prompt = f"""You are SheetCopilot in the REVISING stage.

**Original Task**: {instruction}
**Target Position**: {answer_position}
**Input File**: {file_path}
**Output File**: {output_path}

**Your Previous Plan**:
{proposal.get('plan_extracted', 'No plan extracted')}

**Your Previous Code**:
```python
{proposal['code']}
```

**Execution Result (Contains Errors)**:
{execution_result}

**Your Task**:
1. Carefully analyze the error message and traceback
2. Identify the root cause (common issues: wrong cell reference, missing imports, incorrect data type, index out of range, etc.)
3. Fix the code completely
4. Ensure the corrected code is COMPLETE and can run independently

**Common Error Patterns**:
- AttributeError: Check if object/cell exists before accessing
- IndexError: Verify row/column indices are within range
- TypeError: Ensure correct data types (convert str to int/float if needed)
- NameError: Import all required libraries
- KeyError: Check if dictionary key exists
- Formula errors: Use string formulas correctly (e.g., "=SUM(A1:A10)")

**Output Format**:
## Error Analysis
[What went wrong and why]

## Revision Strategy
[How you will fix it - be specific]

## Corrected Code
```python
import openpyxl
# Add any other necessary imports

# Your COMPLETE corrected implementation here
# Make sure it includes:
# 1. Loading the file
# 2. All necessary operations
# 3. Saving the output
# 4. Closing the workbook
```

Provide your complete revision now:
"""
        
        self.logger.debug(f"[REVISING PROMPT]\n{revising_prompt}")
        
        # Get LLM response
        messages = [
            observation['prompt'], 
            observation['response'],
            proposal['prompt'],
            proposal['response'],
            revising_prompt
        ]
        response = get_llm_response(messages, self.opt)
        self.logger.debug(f"[REVISING RESPONSE]\n{response}")
        
        # Extract revised code
        revised_code = extract_code(response)
        self.logger.debug(f"[REVISING CODE]\n{revised_code}")
        
        return {
            'needs_revision': True,
            'prompt': revising_prompt,
            'response': response,
            'revised_code': revised_code,
            'error_analysis': self._extract_error_analysis(response)
        }
    
    def stage_4_executing(self, code: str, max_retries: int = 3) -> Tuple[str, bool]:
        """
        Stage 4: Executing - Execute the code with retry mechanism
        Returns: (execution_result, success_flag)
        """
        self.current_stage = "EXECUTING"
        self.log_stage("EXECUTING", f"Executing code with max {max_retries} retries")
        
        for attempt in range(max_retries):
            self.logger.info(f"[EXECUTING] Attempt {attempt + 1}/{max_retries}")
            
            try:
                result = exec_code(self.client, code)
                self.logger.info(f"[EXECUTING RESULT]\n{result}")
                
                # Check for errors
                if 'Error' not in result and 'Traceback' not in result:
                    self.logger.info(f"[EXECUTING] SUCCESS on attempt {attempt + 1}")
                    return result, True
                else:
                    self.logger.warning(f"[EXECUTING] Error on attempt {attempt + 1}:\n{result}")
                    
            except Exception as e:
                self.logger.error(f"[EXECUTING] Exception on attempt {attempt + 1}: {str(e)}")
                result = f"Exception: {str(e)}"
            
            if attempt < max_retries - 1:
                self.logger.info(f"[EXECUTING] Retrying...")
        
        self.logger.error(f"[EXECUTING] FAILED after {max_retries} attempts")
        return result, False
    
    def _extract_plan(self, response: str) -> str:
        """Extract plan section from response"""
        if "## Plan" in response:
            plan_start = response.find("## Plan")
            plan_end = response.find("```python", plan_start)
            if plan_end == -1:
                plan_end = len(response)
            return response[plan_start:plan_end].strip()
        return "No plan section found"
    
    def _extract_error_analysis(self, response: str) -> str:
        """Extract error analysis from revision response"""
        if "## Error Analysis" in response:
            start = response.find("## Error Analysis")
            end = response.find("##", start + 1)
            if end == -1:
                end = response.find("```python", start)
            if end == -1:
                end = len(response)
            return response[start:end].strip()
        return "No error analysis found"
    
    def solve_task(self, data: Dict[str, Any], file_path: str, output_path: str) -> Dict[str, Any]:
        """
        Main solving pipeline: Observing → Proposing → Executing → Revising (if needed)
        """
        task_id = data['id']
        instruction = data['instruction']
        answer_position = data['answer_position']
        instruction_type = data['instruction_type']
        
        self.logger.info(f"\n\n{'#'*100}\n# Starting Task: {task_id}\n{'#'*100}\n")
        
        conversation = []
        
        try:
            # Stage 1: Observing
            observation = self.stage_1_observing(
                file_path, instruction, answer_position, instruction_type
            )
            conversation.extend([observation['prompt'], observation['response'], observation['result']])
            
            if not observation['success']:
                self.logger.error(f"[TASK {task_id}] Observation failed, aborting task")
                return self._create_failure_result(task_id, instruction_type, conversation, "Observation failed")
            
            # Stage 2: Proposing
            proposal = self.stage_2_proposing(
                observation, file_path, output_path, instruction, answer_position, instruction_type
            )
            conversation.extend([proposal['prompt'], proposal['response']])
            
            # Stage 4: First Execution
            exec_result, success = self.stage_4_executing(proposal['code'])
            conversation.append(exec_result)
            
            # Stage 3: Revising (if needed)
            max_revisions = self.opt.max_revisions
            revision_count = 0
            
            while not success and revision_count < max_revisions:
                revision_count += 1
                self.logger.info(f"[TASK {task_id}] Revision round {revision_count}/{max_revisions}")
                
                revision = self.stage_3_revising(
                    observation, proposal, exec_result, 
                    file_path, output_path, instruction, answer_position
                )
                
                if not revision['needs_revision']:
                    break
                
                conversation.extend([revision['prompt'], revision['response']])
                
                # Execute revised code
                exec_result, success = self.stage_4_executing(revision['revised_code'])
                conversation.append(exec_result)
                
                # Update proposal with revised code for next iteration
                proposal['code'] = revision['revised_code']
            
            # Check final output file
            output_exists = self._check_output_exists(output_path)
            
            result = {
                'id': task_id,
                'instruction_type': instruction_type,
                'conversation': conversation,
                'solution': proposal['code'],
                'success': success and output_exists,
                'revision_count': revision_count,
                'stage_history': self.stage_history.copy()
            }
            
            self.logger.info(f"[TASK {task_id}] {'SUCCESS' if result['success'] else 'FAILED'} "
                           f"(revisions: {revision_count})")
            
            # Clear stage history for next task
            self.stage_history = []
            
            return result
            
        except Exception as e:
            self.logger.error(f"[TASK {task_id}] Unexpected error: {str(e)}", exc_info=True)
            return self._create_failure_result(task_id, instruction_type, conversation, str(e))
    
    def _check_output_exists(self, output_path: str) -> bool:
        """Check if output file was created"""
        local_path = output_path.replace('/mnt/data', f'../data')
        exists = os.path.exists(local_path)
        self.logger.debug(f"[OUTPUT CHECK] {local_path} exists: {exists}")
        return exists
    
    def _create_failure_result(self, task_id, instruction_type, conversation, error_msg):
        """Create a failure result object"""
        return {
            'id': task_id,
            'instruction_type': instruction_type,
            'conversation': conversation,
            'solution': '',
            'success': False,
            'error': error_msg,
            'stage_history': self.stage_history.copy()
        }


def gen_solution(opt):
    """Main generation function"""
    # Setup logger
    logger = setup_logger('log', opt.model)
    logger.info(f"Starting SheetCopilot with config: {vars(opt)}")
    
    # Load dataset
    dataset_path = os.path.abspath(f'../data/{opt.dataset}')
    with open(f'{dataset_path}/dataset.json', 'r') as fp:
        dataset = json.load(fp)
    
    logger.info(f"Loaded {len(dataset)} tasks from {opt.dataset}")
    
    # Create output directories
    output_dir = f'{dataset_path}/outputs'
    os.makedirs(output_dir, exist_ok=True)
    
    model_output_dir = f'{output_dir}/sheetcopilot_{opt.model}'
    os.makedirs(model_output_dir, exist_ok=True)
    
    # Initialize SheetCopilot
    copilot = SheetCopilot(opt, logger)
    
    # Process each task
    results = []
    success_count = 0
    
    for data in tqdm(dataset, desc="Processing tasks"):
        file_name = f"1_{data['spreadsheet_path'].lstrip('spreadsheet/')}_input.xlsx"
        file_path = f"/mnt/data/{opt.dataset}/{data['spreadsheet_path']}/{file_name}"
        output_path = f"/mnt/data/{opt.dataset}/outputs/sheetcopilot_{opt.model}/{file_name.replace('_input', '_output')}"
        
        # Solve task
        result = copilot.solve_task(data, file_path, output_path)
        results.append(result)
        
        if result['success']:
            success_count += 1
        
        # Save incremental results
        with open(f'{output_dir}/conv_sheetcopilot_{opt.model}.jsonl', 'a', encoding='utf-8') as fp:
            fp.write(json.dumps(result, ensure_ascii=False) + '\n')
    
    # Final statistics
    total_tasks = len(dataset)
    success_rate = (success_count / total_tasks) * 100 if total_tasks > 0 else 0
    
    logger.info(f"\n{'='*80}")
    logger.info(f"FINAL STATISTICS")
    logger.info(f"{'='*80}")
    logger.info(f"Total tasks: {total_tasks}")
    logger.info(f"Successful: {success_count}")
    logger.info(f"Failed: {total_tasks - success_count}")
    logger.info(f"Success rate: {success_rate:.2f}%")
    logger.info(f"{'='*80}\n")
    
    # Save summary
    summary = {
        'model': opt.model,
        'dataset': opt.dataset,
        'total_tasks': total_tasks,
        'successful': success_count,
        'failed': total_tasks - success_count,
        'success_rate': success_rate,
        'config': vars(opt)
    }
    
    with open(f'{output_dir}/summary_sheetcopilot_{opt.model}.json', 'w', encoding='utf-8') as fp:
        json.dump(summary, fp, ensure_ascii=False, indent=2)
    
    return results


def run_solution(opt):
    """Apply solutions to test cases 2 and 3"""
    logger = setup_logger('log', opt.model)
    client = get_exec_client(opt.code_exec_url, opt.conv_id)
    
    dataset_path = os.path.abspath(f'../data/{opt.dataset}')
    conv_file = f'{dataset_path}/outputs/conv_sheetcopilot_{opt.model}.jsonl'
    
    if not os.path.exists(conv_file):
        logger.error(f"Conversation file not found: {conv_file}")
        return
    
    with open(conv_file, 'r', encoding='utf-8') as fp:
        conv_records = [json.loads(line) for line in fp.readlines() if line.strip()]
    
    logger.info(f"Applying solutions to test cases 2 and 3 for {len(conv_records)} tasks")
    
    for conv in tqdm(conv_records, desc="Applying solutions"):
        if not conv.get('success', False):
            logger.warning(f"Skipping failed task: {conv['id']}")
            continue
        
        try:
            for idx in range(2, 4):
                input_file = f"{idx}_{conv['id']}_input.xlsx"
                output_file = f"{idx}_{conv['id']}_output.xlsx"
                
                # Replace file names in solution
                solution = conv['solution'].replace(f"1_{conv['id']}_input.xlsx", input_file)
                solution = solution.replace(f"1_{conv['id']}_output.xlsx", output_file)
                
                logger.info(f"Applying solution to {conv['id']} test case {idx}")
                exec_result = exec_code(client, solution)
                
                if 'Error' in exec_result:
                    logger.warning(f"Error in test case {idx} for {conv['id']}: {exec_result}")
                else:
                    logger.info(f"Successfully applied to test case {idx} for {conv['id']}")
                    
        except Exception as e:
            logger.error(f"Error processing {conv['id']}: {str(e)}")


def parse_option():
    parser = argparse.ArgumentParser("SheetCopilot: Multi-stage reasoning for spreadsheet manipulation")
    
    parser.add_argument('--model', type=str, required=True, help='Model name')
    parser.add_argument('--api_key', type=str, default="", help='API key')
    parser.add_argument('--base_url', type=str, default="", help='Base URL')
    parser.add_argument('--dataset', type=str, default="test1", help='Dataset name')
    parser.add_argument('--code_exec_url', type=str, default="http://localhost:8080/execute", 
                       help='Code execution URL')
    parser.add_argument('--conv_id', type=str, default="EVAL", help='Conversation ID')
    parser.add_argument('--max_revisions', type=int, default=3, 
                       help='Maximum number of revision rounds')
    parser.add_argument('--skip_run_solution', default=False, action='store_true',
                       help='Skip running solution on test cases 2 and 3')
    
    opt = parser.parse_args()
    return opt


if __name__ == '__main__':
    opt = parse_option()
    
    # Run generation
    results = gen_solution(opt)
    
    # Run on test cases if not skipped
    if not opt.skip_run_solution:
        run_solution(opt)
