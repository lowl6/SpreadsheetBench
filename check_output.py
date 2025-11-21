import openpyxl
from pathlib import Path

# 检查 test1 数据集的 57072 输出
output_path = Path('data/test1/outputs/sheetcopilot_glm-4.5-air/1_57072_output.xlsx')

if not output_path.exists():
    print(f'❌ File not found: {output_path}')
else:
    wb = openpyxl.load_workbook(output_path, data_only=True)
    ws = wb['Sheet2'] if 'Sheet2' in wb.sheetnames else wb.active
    print(f'✅ File exists, using sheet: {ws.title}')
    
    # 检查 B 列前 120 行
    non_empty = []
    for row in range(1, 121):
        cell = ws[f'B{row}']
        if cell.value is not None and cell.value != "":
            non_empty.append((cell.coordinate, cell.value))
    
    print(f'Non-empty cells in B1:B120: {len(non_empty)}')
    if len(non_empty) > 0:
        print(f'First 15: {non_empty[:15]}')
        print(f'\n✅ SUCCESS! Excel calculated the formulas.')
    else:
        print(f'\n❌ FAILED! B column is still empty.')
        # 检查是否还有公式
        ws_formula = openpyxl.load_workbook(output_path, data_only=False)['Sheet2']
        b1_formula = ws_formula['B1'].value
        print(f'B1 content (data_only=False): {b1_formula}')
    
    wb.close()
